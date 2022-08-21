import openpyxl as xls

table = []

# 1 столбец
def left_formula(y):
    global table

    res = ["-"]
    for i in range(1, len(y)):
        r = y[i] - y[i - 1]
        res.append(str(round(r, 3)))

    table.append(res)


# 2 столбец
def center_formula(y):
    global table

    res = ["-"]
    for i in range(1, len(y) - 1):
        r = (y[i + 1] - y[i - 1]) / 2
        res.append(str(round(r, 3)))
    res.append("-")

    table.append(res)


# 3 столбец
def equal_param_right_formula(y):
    global table

    res = []
    for i in range(len(y) - 2):
        s1 = y[i + 1] - y[i]
        s2 = (y[i + 2] - y[i]) / 2
        r = s1 * 2 - s2
        res.append(str(round(r, 3)))
    res.append("-")
    res.append("-")

    table.append(res)
    

# 4 столбец
def var_der_formula(x, y):
    global table

    res = []
    for i in range(len(y) - 1):
        s1 = (1 / y[i] - 1 / y[i + 1]) / (1 / x[i] - 1 / x[i + 1])
        r = y[i] * y[i] * s1 / (x[i] * x[i])
        res.append(str(round(r, 3)))
    res.append("-")
    res.append("-")

    table.append(res)


# 5 столбец
def second_dif_formula(y):
    global table

    res = ["-"]
    for i in range(1, len(y) - 1):
        r = y[i - 1] - 2 * y[i] + y[i + 1]
        res.append(str(round(r, 3)))
    res.append("-")

    table.append(res)


def parse_table(name):
    """
        Загрузка таблицы в программу.
    """
    try:
        pos = 2
        points = xls.load_workbook(name).active
        table = []
        while points.cell(row = pos, column = 1).value is not None:
            table.append([float(points.cell(row = pos, column = 1).value), \
                          float(points.cell(row = pos, column = 2).value)])
            pos += 1

        res = [[],[]]
        for i in range(0, len(table)):
            res[0].append(table[i][0])
            res[1].append(table[i][1])
        return res
        
    except TypeError:
        print("Проверьте данные на вводе!")
        return None, None, None
    except ValueError:
        print("Проверьте данные на вводе!!!")
        return None, None, None


def print_table(table):
    print("\n\n|  x  |   y   |   1   |   2   |   3   |   4   |   5   |")
    print("-------------------------------------------------------")
    for i in range(len(table[0])):
        print("|%5s|%7s|%7s|%7s|%7s|%7s|%7s|"%(table[0][i], table[1][i], 
                                               table[2][i], table[3][i], 
                                               table[4][i], table[5][i], 
                                               table[6][i]))
    print("\n\n")


def main():
    global table
    table = [[1, 2, 3, 4, 5, 6], [0.571, 0.889, 1.091, 1.231, 1.333, 1.412]]

    # 1 столбец -  левосторонняя формула
    left_formula(table[1])

    # 2 столбец - центральная формула
    center_formula(table[1])

    # 3 столбец - вторая формула Рунге с правосторонней формулы
    equal_param_right_formula(table[1])

    # 4 столбец - метод выравнивающих переменных
    var_der_formula(table[0], table[1])

    # 5 столбец - вторая разностная производная
    second_dif_formula(table[1])

    for i in range(6):
        table[0][i] = str(table[0][i])
        table[1][i] = str(table[1][i])

    print_table(table)
    
if __name__ == "__main__":
    main()